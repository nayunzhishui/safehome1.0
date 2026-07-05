"""Generate offline audit artifacts for every current profile model.

The script reads the current model JSON files, including user-edited
profile names and explanations, and existing derived feature matrices. It does
not rerun clustering and does not rewrite model JSON.
"""

from __future__ import annotations

import json
import math
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from statistics import multimode

import matplotlib.pyplot as plt
from matplotlib import font_manager
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OTHER_ROOT = Path(r"D:\codex\workspace\safehome1.0其他内容")
DESIGN_DIR = OTHER_ROOT / "画像系统设计_Claude_20260628"
PRODUCT_PROFILE_DIR = PROJECT_ROOT / "content" / "profiles"
PRODUCT_MATRIX_DIR = DESIGN_DIR / "03_画像模型派生文件_20260629" / "特征矩阵"
SUPPLEMENT_MODEL_DIR = DESIGN_DIR / "04_未映射问卷补充聚类_20260629" / "模型JSON"
SUPPLEMENT_ROOT = DESIGN_DIR / "04_未映射问卷补充聚类_20260629"
OUTPUT_ROOT = DESIGN_DIR / "05_画像模型离线审核产物_20260629"

sns.set_theme(style="whitegrid")


def configure_chinese_font() -> None:
    font_candidates = [
        Path(r"C:\Windows\Fonts\NotoSansSC-VF.ttf"),
        Path(r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simhei.ttf"),
        Path(r"C:\Windows\Fonts\simsun.ttc"),
    ]
    fallback_fonts = ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Arial Unicode MS"]
    for font_path in font_candidates:
        if font_path.exists():
            font_manager.fontManager.addfont(str(font_path))
            font_name = font_manager.FontProperties(fname=str(font_path)).get_name()
            fallback_fonts.insert(0, font_name)
            break
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = fallback_fonts
    plt.rcParams["axes.unicode_minus"] = False


configure_chinese_font()


@dataclass
class AuditModel:
    model_type: str
    source_json: Path
    payload: dict
    matrix_path: Path


def safe_slug(value: str, max_length: int = 90) -> str:
    slug = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "_", str(value)).strip("_")
    return (slug[:max_length] or "model").strip("_")


def short_label(value: str, max_length: int = 24) -> str:
    text = re.sub(r"^\d+[.、，,]?\s*", "", str(value or "")).strip()
    return text[:max_length] + ("…" if len(text) > max_length else "")


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, encoding="utf-8-sig")


def feature_ids(model: dict) -> list[str]:
    return [str(item.get("feature_id")) for item in model.get("features", []) if item.get("feature_id")]


def matrix_feature_columns(matrix: pd.DataFrame, model: dict) -> list[str]:
    ids = feature_ids(model)
    return [column for column in ids if column in matrix.columns]


def cluster_counts_match(matrix: pd.DataFrame, model: dict) -> bool:
    if "cluster_id" not in matrix.columns:
        return False
    expected = {int(cluster.get("cluster_id")): int(cluster.get("n")) for cluster in model.get("clusters", [])}
    actual = {int(key): int(value) for key, value in matrix["cluster_id"].value_counts().to_dict().items()}
    return expected == actual


def center_diff(matrix: pd.DataFrame, model: dict, columns: list[str]) -> float:
    if not columns or "cluster_id" not in matrix.columns:
        return float("inf")
    means = {item.get("feature_id"): item.get("mean") for item in model.get("features", [])}
    stds = {item.get("feature_id"): item.get("std") for item in model.get("features", [])}
    total = 0.0
    count = 0
    for cluster in model.get("clusters", []):
        cluster_id = cluster.get("cluster_id")
        rows = matrix[matrix["cluster_id"] == cluster_id]
        if rows.empty:
            continue
        expected_center = cluster.get("center_z") or {}
        for column in columns:
            mean = means.get(column)
            std = stds.get(column) or 1
            if mean is None:
                continue
            observed = (pd.to_numeric(rows[column], errors="coerce").mean() - float(mean)) / float(std)
            expected = expected_center.get(column)
            if expected is None or math.isnan(observed):
                continue
            total += abs(observed - float(expected))
            count += 1
    return total / count if count else float("inf")


def find_product_matrix(model: dict) -> Path:
    ids = set(feature_ids(model))
    candidates: list[tuple[float, Path]] = []
    for path in PRODUCT_MATRIX_DIR.glob("*.csv"):
        try:
            matrix = read_csv(path)
        except Exception:
            continue
        if len(matrix) != int(model.get("n_cases", -1)):
            continue
        if not ids.issubset(set(matrix.columns)):
            continue
        if not cluster_counts_match(matrix, model):
            continue
        diff = center_diff(matrix, model, list(ids))
        candidates.append((diff, path))
    if not candidates:
        raise FileNotFoundError(f"未找到匹配的产品候选特征矩阵：{model.get('model_id')}")
    return sorted(candidates, key=lambda item: item[0])[0][1]


def load_models() -> list[AuditModel]:
    models: list[AuditModel] = []
    for path in sorted(PRODUCT_PROFILE_DIR.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        if payload.get("schema_version") != "2026.06-profile-model-v1":
            continue
        matrix_path = find_product_matrix(payload)
        models.append(AuditModel("产品候选画像模型", path, payload, matrix_path))
    for path in sorted(SUPPLEMENT_MODEL_DIR.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        if payload.get("schema_version") != "2026.06-unmapped-supplementary-profile-v1":
            continue
        matrix_path = SUPPLEMENT_ROOT / payload["matrix_file"]
        models.append(AuditModel("未映射补充画像模型", path, payload, matrix_path))
    return models


def label_lookup(model: dict) -> dict[str, str]:
    return {
        str(item.get("feature_id")): str(item.get("label") or item.get("feature_id"))
        for item in model.get("features", [])
    }


def cluster_summary_rows(model: dict) -> list[dict]:
    rows = []
    for cluster in model.get("clusters", []):
        rows.append(
            {
                "cluster_id": cluster.get("cluster_id"),
                "画像名称": cluster.get("profile_name"),
                "样本数": cluster.get("n"),
                "占比": cluster.get("percent"),
                "支持性解释": cluster.get("supportive_explanation"),
            }
        )
    return rows


def mode_value(series: pd.Series):
    values = [item for item in series.dropna().tolist()]
    if not values:
        return None
    try:
        return multimode(values)[0]
    except Exception:
        return values[0]


def feature_descriptives(matrix: pd.DataFrame, columns: list[str], labels: dict[str, str]) -> pd.DataFrame:
    rows = []
    for column in columns:
        values = pd.to_numeric(matrix[column], errors="coerce")
        rows.append(
            {
                "变量ID": column,
                "题目/维度标签": labels.get(column, column),
                "mean": round(float(values.mean()), 4),
                "std": round(float(values.std()), 4),
                "median": round(float(values.median()), 4),
                "mode": mode_value(values),
                "min": round(float(values.min()), 4),
                "max": round(float(values.max()), 4),
                "non_missing": int(values.notna().sum()),
            }
        )
    return pd.DataFrame(rows)


def cluster_feature_means(matrix: pd.DataFrame, columns: list[str], labels: dict[str, str]) -> pd.DataFrame:
    mean_df = matrix.groupby("cluster_id")[columns].mean().round(4)
    mean_df = mean_df.rename(columns={column: short_label(labels.get(column, column), 32) for column in columns})
    mean_df.index.name = "cluster_id"
    return mean_df.reset_index()


def cluster_feature_stats(matrix: pd.DataFrame, columns: list[str], labels: dict[str, str]) -> pd.DataFrame:
    rows = []
    for cluster_id, group in matrix.groupby("cluster_id"):
        for column in columns:
            values = pd.to_numeric(group[column], errors="coerce")
            rows.append(
                {
                    "cluster_id": cluster_id,
                    "变量ID": column,
                    "题目/维度标签": labels.get(column, column),
                    "mean": round(float(values.mean()), 4),
                    "std": round(float(values.std()), 4),
                    "median": round(float(values.median()), 4),
                    "mode": mode_value(values),
                    "n": int(values.notna().sum()),
                }
            )
    return pd.DataFrame(rows)


def model_selection_df(model: dict) -> pd.DataFrame:
    rows = []
    for item in model.get("model_selection", []):
        rows.append(
            {
                "k": item.get("k"),
                "inertia": item.get("inertia"),
                "silhouette": item.get("silhouette"),
                "min_cluster_size": item.get("min_cluster_size"),
            }
        )
    return pd.DataFrame(rows)


def feature_meta_df(model: dict) -> pd.DataFrame:
    rows = []
    for item in model.get("features", []):
        rows.append(
            {
                "feature_id": item.get("feature_id"),
                "source_variable": item.get("source_variable"),
                "worksheet_question_id": item.get("worksheet_question_id"),
                "question_no": item.get("question_no"),
                "label": item.get("label"),
                "reverse_scored": item.get("reverse_scored"),
                "mean": item.get("mean"),
                "std": item.get("std"),
            }
        )
    return pd.DataFrame(rows)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="E8EEF5")
    title_fill = PatternFill("solid", fgColor="DDEFEA")
    thin = Side(style="thin", color="D9E2E7")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = title_fill
                cell.font = Font(bold=True, color="173B35")
            elif cell.row == 2:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            cell.border = border
    for column_cells in ws.columns:
        max_len = 8
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells[:80]:
            max_len = max(max_len, min(len(str(cell.value or "")), 48))
        ws.column_dimensions[column].width = min(max_len + 2, 42)
    ws.freeze_panes = "A3"


def write_df(ws, title: str, df: pd.DataFrame) -> None:
    ws.append([title])
    if df.empty:
        ws.append(["无数据"])
        return
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    style_sheet(ws)


def write_workbook(path: Path, model: dict, matrix_path: Path, tables: dict[str, pd.DataFrame]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    overview = wb.create_sheet("模型总览")
    overview_df = pd.DataFrame(
        [
            ["模型名称", model.get("standard_scale_name") or model.get("display_name")],
            ["模型ID", model.get("model_id")],
            ["来源数据", model.get("source_dataset")],
            ["样本量", model.get("n_cases")],
            ["特征数", model.get("n_features")],
            ["选择k", model.get("chosen_k")],
            ["边界说明", model.get("boundary_notice")],
            ["特征矩阵", str(matrix_path)],
        ],
        columns=["字段", "内容"],
    )
    write_df(overview, "模型总览", overview_df)
    for sheet_name, df in tables.items():
        ws = wb.create_sheet(sheet_name[:31])
        write_df(ws, sheet_name, df)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def plot_k_selection(path: Path, model: dict) -> None:
    df = model_selection_df(model)
    if df.empty:
        return
    fig, axes = plt.subplots(1, 2, figsize=(11, 4.5))
    axes[0].plot(df["k"], df["inertia"], marker="o", color="#287271")
    axes[0].set_title("Elbow Method")
    axes[0].set_xlabel("k")
    axes[0].set_ylabel("Inertia")
    axes[1].plot(df["k"], df["silhouette"], marker="o", color="#C75D2C")
    axes[1].set_title("Silhouette Score")
    axes[1].set_xlabel("k")
    axes[1].set_ylabel("score")
    for ax in axes:
        ax.axvline(model.get("chosen_k"), linestyle="--", color="#6C757D", linewidth=1)
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)


def plot_pca(path: Path, model: dict, matrix: pd.DataFrame) -> None:
    if not {"pc1", "pc2", "cluster_id"}.issubset(matrix.columns):
        return
    fig, ax = plt.subplots(figsize=(7.5, 6))
    sns.scatterplot(
        data=matrix,
        x="pc1",
        y="pc2",
        hue="cluster_id",
        palette="Set2",
        s=34,
        alpha=0.72,
        ax=ax,
    )
    for cluster in model.get("clusters", []):
        centroid = cluster.get("pca_centroid") or {}
        if centroid.get("pc1") is None or centroid.get("pc2") is None:
            continue
        ax.scatter(centroid["pc1"], centroid["pc2"], s=150, marker="X", color="#102A43", edgecolor="white", linewidth=1.5)
        ax.text(centroid["pc1"], centroid["pc2"], f"画像{int(cluster.get('cluster_id')) + 1}", fontsize=9)
    ax.set_title("聚类分布（PCA 二维投影）")
    ax.set_xlabel("PC1")
    ax.set_ylabel("PC2")
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)


def plot_heatmap(path: Path, means: pd.DataFrame) -> None:
    if means.empty:
        return
    heat = means.set_index("cluster_id")
    width = max(8, min(18, 0.45 * len(heat.columns) + 4))
    fig, ax = plt.subplots(figsize=(width, 4.8))
    sns.heatmap(heat, annot=True, cmap="YlGnBu", fmt=".2f", linewidths=0.5, ax=ax)
    ax.set_title("各聚类画像热图（原始特征均值）")
    ax.set_xlabel("特征")
    ax.set_ylabel("cluster")
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)


def plot_lines(path: Path, means: pd.DataFrame) -> None:
    if means.empty:
        return
    line_df = means.set_index("cluster_id").T
    fig, ax = plt.subplots(figsize=(max(9, min(20, 0.45 * len(line_df.index) + 4)), 5.5))
    line_df.plot(marker="o", ax=ax)
    ax.set_title("各聚类画像折线图（原始特征均值）")
    ax.set_xlabel("特征")
    ax.set_ylabel("均值")
    ax.tick_params(axis="x", rotation=60)
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)


def write_cluster_summary(path: Path, model: dict) -> None:
    lines = [
        "各聚类类别样本分布",
        "=" * 30,
        f"模型：{model.get('standard_scale_name') or model.get('display_name')}",
        f"来源：{model.get('source_dataset')}",
        "",
    ]
    for cluster in model.get("clusters", []):
        lines.append(f"Cluster {cluster.get('cluster_id')} | {cluster.get('profile_name')} | n={cluster.get('n')} | 占比={cluster.get('percent')}%")
        lines.append(f"解释：{cluster.get('supportive_explanation')}")
    lines.append("")
    lines.append(f"总样本数：{model.get('n_cases')}")
    path.write_text("\n".join(lines), encoding="utf-8")


def process_model(index: int, audit_model: AuditModel, output_group: Path) -> dict:
    model = audit_model.payload
    matrix = read_csv(audit_model.matrix_path)
    columns = matrix_feature_columns(matrix, model)
    labels = label_lookup(model)
    model_name = model.get("standard_scale_name") or model.get("display_name") or model.get("model_id")
    folder_name = f"{index:02d}_{safe_slug(model_name, 36)}_{safe_slug(model.get('source_dataset', ''), 42)}"
    model_dir = output_group / folder_name
    model_dir.mkdir(parents=True, exist_ok=True)

    cluster_df = pd.DataFrame(cluster_summary_rows(model))
    desc_df = feature_descriptives(matrix, columns, labels)
    means_df = cluster_feature_means(matrix, columns, labels)
    stats_df = cluster_feature_stats(matrix, columns, labels)
    selection_df = model_selection_df(model)
    features_df = feature_meta_df(model)

    tables = {
        "cluster_summary": cluster_df,
        "feature_descriptives": desc_df,
        "cluster_feature_means": means_df,
        "cluster_feature_stats": stats_df,
        "model_selection": selection_df,
        "features": features_df,
    }
    write_workbook(model_dir / "画像模型离线审核表.xlsx", model, audit_model.matrix_path, tables)
    for name, df in tables.items():
        df.to_csv(model_dir / f"{name}.csv", index=False, encoding="utf-8-sig")

    write_cluster_summary(model_dir / "cluster_summary.txt", model)
    plot_k_selection(model_dir / "K-means过程.png", model)
    plot_pca(model_dir / "问卷聚类分布_PCA.png", model, matrix)
    plot_heatmap(model_dir / "各聚类画像热图_原始特征.png", means_df)
    plot_lines(model_dir / "各聚类画像折线图_原始特征.png", means_df)
    shutil.copy2(audit_model.source_json, model_dir / "来源模型.json")

    return {
        "序号": index,
        "模型类型": audit_model.model_type,
        "模型名称": model_name,
        "来源数据": model.get("source_dataset"),
        "样本量": model.get("n_cases"),
        "特征数": model.get("n_features"),
        "k": model.get("chosen_k"),
        "输出目录": str(model_dir),
        "审核表": str(model_dir / "画像模型离线审核表.xlsx"),
        "PCA图": str(model_dir / "问卷聚类分布_PCA.png"),
        "热图": str(model_dir / "各聚类画像热图_原始特征.png"),
    }


def write_index(rows: list[dict]) -> None:
    index_df = pd.DataFrame(rows)
    index_df.to_csv(OUTPUT_ROOT / "画像模型离线审核产物索引.csv", index=False, encoding="utf-8-sig")
    wb = Workbook()
    ws = wb.active
    ws.title = "产物索引"
    write_df(ws, "画像模型离线审核产物索引", index_df)
    wb.save(OUTPUT_ROOT / "画像模型离线审核产物索引.xlsx")

    lines = [
        "# 画像模型离线审核产物索引",
        "",
        "更新日期：2026-06-29",
        "",
        "本目录为画像模型离线人工审核产物，仿照 `LPAandKmeans` 的产物格式补充 PNG、Excel、CSV 和摘要文本。",
        "",
        "重要边界：本脚本只读取当前模型 JSON 和已有特征矩阵，不重跑聚类；`content/profiles` 中用户已修改的画像命名和解释会原样保留。",
        "",
        "## 产物范围",
        "",
        f"- 产品候选画像模型：{sum(1 for row in rows if row['模型类型'] == '产品候选画像模型')} 个",
        f"- 未映射补充画像模型：{sum(1 for row in rows if row['模型类型'] == '未映射补充画像模型')} 个",
        "",
        "每个模型目录包含：",
        "",
        "- `画像模型离线审核表.xlsx`",
        "- `cluster_summary.txt`",
        "- `K-means过程.png`",
        "- `问卷聚类分布_PCA.png`",
        "- `各聚类画像热图_原始特征.png`",
        "- `各聚类画像折线图_原始特征.png`",
        "- `cluster_summary.csv`、`feature_descriptives.csv`、`cluster_feature_means.csv`、`cluster_feature_stats.csv`、`model_selection.csv`、`features.csv`",
        "",
        "## 模型清单",
        "",
        "| 序号 | 类型 | 模型 | n | 特征数 | k | 输出目录 |",
        "| ---: | --- | --- | ---: | ---: | ---: | --- |",
    ]
    for row in rows:
        rel = Path(row["输出目录"]).relative_to(OUTPUT_ROOT)
        lines.append(f"| {row['序号']} | {row['模型类型']} | {row['模型名称']} | {row['样本量']} | {row['特征数']} | {row['k']} | `{rel}` |")
    lines.append("")
    (OUTPUT_ROOT / "画像模型离线审核产物索引.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    product_out = OUTPUT_ROOT / "产品候选画像模型"
    supplement_out = OUTPUT_ROOT / "未映射补充画像模型"
    product_out.mkdir(parents=True, exist_ok=True)
    supplement_out.mkdir(parents=True, exist_ok=True)

    rows: list[dict] = []
    product_index = 0
    supplement_index = 0
    for audit_model in load_models():
        if audit_model.model_type == "产品候选画像模型":
            product_index += 1
            rows.append(process_model(product_index, audit_model, product_out))
        else:
            supplement_index += 1
            rows.append(process_model(supplement_index, audit_model, supplement_out))
    write_index(rows)
    print(json.dumps({"generated_models": len(rows), "output_root": str(OUTPUT_ROOT)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
