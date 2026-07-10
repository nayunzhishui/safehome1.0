"""Prepare Task 12 relationship item matrices without exporting row-level data.

The script reads the user-provided workbooks outside the repository, builds an
auditable item crosswalk, recomputes frozen dimensions, and writes only safe
aggregate artifacts plus an ignored local NPZ used by the profiling step.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = Path(r"D:\codex\workspace\safehome1.0其他内容\夏老师文件\数据1")
ORIGINAL_PATH = DATA_ROOT / "原始量表.xlsx"
ALL_DATA_PATH = DATA_ROOT / "全部数据1.0.xlsx"
CLEANED_PATH = DATA_ROOT / "清洗好的469份.xlsx"
SCORING_PATH = DATA_ROOT / "问卷计分算分说明书.docx"
OUTPUT_ROOT = PROJECT_ROOT / "outputs" / "task12_relationship_profiles"
PRIVATE_ROOT = OUTPUT_ROOT / "private"
DOC_OUTPUT = PROJECT_ROOT / "docs" / "02_专项进度与验收" / "任务十二数据1题项缩写映射表.md"

RF_COLUMNS = [f"Q{index}" for index in range(1, 19)]
YSQ_COLUMNS = [f"YSQ{index}" for index in range(1, 19)]
BELIEF_COLUMNS = [name for index in range(1, 6) for name in (f"a{index}", f"b{index}")]
SN_COLUMNS = [f"SN{index}" for index in range(1, 5)]
PBC_COLUMNS = [f"PBC{index}" for index in range(1, 7)]
BI_COLUMNS = [f"BI{index}" for index in range(1, 7)]
RAP_COLUMNS = [f"RAP{index}" for index in range(1, 6)]
RELATIONSHIP_COLUMNS = BELIEF_COLUMNS + SN_COLUMNS + PBC_COLUMNS + BI_COLUMNS + RAP_COLUMNS
CORE_ITEM_COLUMNS = RF_COLUMNS + YSQ_COLUMNS + RELATIONSHIP_COLUMNS
ITEM_COLUMNS = CORE_ITEM_COLUMNS + ["@11", "@12"]

PROMOTION_ITEMS = ["Q3", "Q5", "Q6", "Q8", "Q12", "Q14", "Q16", "Q17", "Q18"]
PREVENTION_ITEMS = ["Q1", "Q2", "Q4", "Q7", "Q9", "Q10", "Q11", "Q13", "Q15"]
REL_SCHEMA_ITEMS = ["YSQ1", "YSQ2", "YSQ3", "YSQ4", "YSQ5", "YSQ8", "YSQ9", "YSQ13", "YSQ14", "YSQ18"]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_excel_frame(path: Path) -> pd.DataFrame:
    """Read values through openpyxl to avoid pandas/openpyxl version coupling."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    data = list(rows)
    workbook.close()
    return pd.DataFrame(data, columns=headers)


def _actual_header(canonical: str, headers: list[str]) -> str:
    if canonical in headers:
        return canonical
    if canonical.startswith("@"):
        match = next((header for header in headers if str(header).startswith(canonical)), None)
        if match:
            return str(match)
    raise ValueError(f"全部数据1.0.xlsx 缺少题项列：{canonical}")


def _mapping_meta(index: int, abbreviation: str) -> tuple[str, str, int, str, bool]:
    if index < 18:
        return "regulatory_focus_relationship_18", "调节聚焦", index + 1, "1-5（本数据实测；新问卷文档写1-7，待人工确认）", True
    if index < 36:
        return "micro_ysq_relationship_18", "Micro YSQ-18", index - 17, "1-6", True
    relation_index = index - 36
    if relation_index < 31:
        return "relationship_initiation_intention_action", "亲密关系启动意向与主动行为", relation_index + 1, "1-5", True
    return "relationship_initiation_intention_action", "开放叙事", relation_index + 1, "文本", False


def build_item_mapping(original_headers: list[object], all_data_headers: list[object]) -> list[dict]:
    """Map 69 abbreviated columns to the 69 complete prompts by frozen order."""
    original = [str(value or "").strip() for value in original_headers]
    all_headers = [str(value or "").strip() for value in all_data_headers]
    if len(original) < 79:
        raise ValueError(f"原始量表.xlsx 列数不足：expected>=79 actual={len(original)}")
    prompt_headers = original[10:79]
    if len(prompt_headers) != len(ITEM_COLUMNS):
        raise ValueError("原始题干数量与缩写列数量不一致")
    mapping: list[dict] = []
    for index, (canonical, prompt) in enumerate(zip(ITEM_COLUMNS, prompt_headers)):
        actual = _actual_header(canonical, all_headers)
        scale_id, scale_name, question_no, score_range, used = _mapping_meta(index, canonical)
        mapping.append(
            {
                "scale_id": scale_id,
                "scale_name": scale_name,
                "question_no": question_no,
                "abbreviation": canonical,
                "source_column": actual,
                "original_prompt": prompt,
                "score_range": score_range,
                "used_for_clustering": used,
            }
        )
    return mapping


def _numeric(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    return frame[columns].apply(pd.to_numeric, errors="coerce")


def calculate_dimensions(frame: pd.DataFrame) -> pd.DataFrame:
    """Recompute dimensions exactly as frozen in Claude Task 12 T12-08."""
    data = _numeric(frame, CORE_ITEM_COLUMNS)
    result = pd.DataFrame(index=data.index)
    result["PROM"] = data[PROMOTION_ITEMS].mean(axis=1)
    result["PREV"] = data[PREVENTION_ITEMS].mean(axis=1)
    result["RFD"] = result["PROM"] - result["PREV"]
    result["EMS_M"] = data[YSQ_COLUMNS].mean(axis=1)
    result["EMS_SUM"] = data[YSQ_COLUMNS].sum(axis=1, min_count=len(YSQ_COLUMNS))
    result["REL_SCHEMA"] = data[REL_SCHEMA_ITEMS].mean(axis=1)
    result["B1"] = data["a1"] * data["b1"]
    result["B2"] = data["a2"] * data["b2"]
    result["B3"] = data["a3"] * data["b3"]
    result["BENEFIT"] = result[["B1", "B2", "B3"]].mean(axis=1)
    result["REJ_THREAT"] = data["a4"] * data["b4"]
    result["AUTH_THREAT"] = data["a5"] * data["b5"]
    result["AUTH_PROTECT"] = pd.concat([6 - data["a5"], data["b5"]], axis=1).mean(axis=1)
    result["THREAT"] = result[["REJ_THREAT", "AUTH_THREAT"]].mean(axis=1)
    result["SN"] = data[SN_COLUMNS].mean(axis=1)
    result["PBC"] = data[PBC_COLUMNS].mean(axis=1)
    result["BI"] = data[BI_COLUMNS].mean(axis=1)
    result["RAP"] = data[RAP_COLUMNS].mean(axis=1)
    return result


def _within_source_validation(recomputed: pd.DataFrame, source: pd.DataFrame) -> dict:
    column_map = {
        "PROM": "促进聚焦",
        "PREV": "防御聚焦",
        "RFD": "调节聚焦优势指数",
        "EMS_SUM": "早期图示总分",
        "BENEFIT": "获益预期",
        "REJ_THREAT": "拒绝或尴尬威胁",
        "AUTH_THREAT": "真实性威胁冲突",
        "THREAT": "威胁预期总分",
        "AUTH_PROTECT": "真实性保护指数",
        "SN": "主观规范计分",
        "PBC": "知觉行为控制",
        "BI": "BI整体行为意向总分",
        "RAP": "关系主动性实践",
    }
    output: dict[str, dict] = {}
    for computed_name, source_name in column_map.items():
        if source_name not in source.columns:
            output[computed_name] = {"source_column": source_name, "available": False}
            continue
        expected = pd.to_numeric(source[source_name], errors="coerce")
        pair = pd.concat([recomputed[computed_name], expected], axis=1).dropna()
        diff = (pair.iloc[:, 0] - pair.iloc[:, 1]).abs()
        output[computed_name] = {
            "source_column": source_name,
            "available": True,
            "n_compared": int(len(diff)),
            "mean_absolute_difference": round(float(diff.mean()), 10) if len(diff) else None,
            "max_absolute_difference": round(float(diff.max()), 10) if len(diff) else None,
            "exact_within_1e_8": bool((diff <= 1e-8).all()) if len(diff) else False,
        }
    return output


def _cross_sample_summary(recomputed: pd.DataFrame, cleaned: pd.DataFrame) -> dict:
    column_map = {
        "PROM": "PF",
        "PREV": "PrF",
        "RFD": "PF_PrF",
        "EMS_SUM": "YSQ",
        "BENEFIT": "BE",
        "REJ_THREAT": "拒绝或尴尬威胁",
        "AUTH_THREAT": "真实性威胁冲突",
        "THREAT": "TE",
        "AUTH_PROTECT": "RP",
        "SN": "SN",
        "PBC": "PBC",
        "BI": "BI",
        "RAP": "RPP",
    }
    output: dict[str, dict] = {}
    for computed_name, cleaned_name in column_map.items():
        if cleaned_name not in cleaned.columns:
            output[computed_name] = {"cleaned_column": cleaned_name, "available": False}
            continue
        left = pd.to_numeric(recomputed[computed_name], errors="coerce")
        right = pd.to_numeric(cleaned[cleaned_name], errors="coerce")
        left_mean = float(left.mean())
        right_mean = float(right.mean())
        output[computed_name] = {
            "cleaned_column": cleaned_name,
            "available": True,
            "recomputed_n": int(left.notna().sum()),
            "cleaned_n": int(right.notna().sum()),
            "recomputed_mean": round(left_mean, 6),
            "cleaned_mean": round(right_mean, 6),
            "absolute_mean_difference": round(abs(left_mean - right_mean), 6),
            "comparison_scope": "aggregate_only_different_samples",
        }
    return output


def _observed_range(frame: pd.DataFrame, columns: list[str]) -> dict:
    values = _numeric(frame, columns).to_numpy(dtype=float)
    return {
        "columns": len(columns),
        "minimum": float(np.nanmin(values)),
        "maximum": float(np.nanmax(values)),
        "missing_cells": int(np.isnan(values).sum()),
    }


def _mapping_markdown(mapping: list[dict], summary: dict) -> str:
    lines = [
        "# 任务十二数据1题项缩写映射表",
        "",
        "生成方式：`python analysis/profiling/build_task12_relationship_dataset.py`。映射以 `原始量表.xlsx` 第 11-79 列与 `全部数据1.0.xlsx` 的 69 个缩写列固定顺序建立；开放题只进入叙事材料，不进入聚类。",
        "",
        "| 量表 | 题号 | 缩写列 | 原题干 | 分值范围 | 用于聚类 |",
        "| --- | ---: | --- | --- | --- | --- |",
    ]
    for item in mapping:
        prompt = str(item["original_prompt"]).replace("|", "\\|").replace("\n", " ")
        lines.append(
            f"| {item['scale_name']} | {item['question_no']} | `{item['source_column']}` | {prompt} | "
            f"{item['score_range']} | {'是' if item['used_for_clustering'] else '否'} |"
        )
    lines.extend(
        [
            "",
            "## 自动核对结论",
            "",
            f"- 原始量表数据行：{summary['sample_counts']['original']}；全部数据1.0 数据行：{summary['sample_counts']['all_data']}；清洗维度表数据行：{summary['sample_counts']['cleaned']}。",
            "- 69 个缩写列已全部映射，其中 67 个量化题项进入分量表聚类，2 个开放题不进入自动聚类。",
            "- `全部数据1.0.xlsx` 与 `清洗好的469份.xlsx` 样本量不同，且清洗表无稳定行 ID，因此只做聚合分布交叉验证，不能伪造逐行对应。",
            "- 调节聚焦题项在当前数据中的观测范围为 1-5，但新版问卷和计分说明写 1-7；画像模型按当前数据标准化训练，新量表开放前需研究者确认展示量尺。",
            "- 当前数据包含 `PBC1-PBC6` 六题，而新版问卷/信度报告按四题 PBC；数据准备按任务十二冻结口径使用六题，前端正式录入前必须人工确认版本。",
            "- `全部数据1.0` 的 PBC 派生列可由 `PBC3-PBC6` 精确复现；其“促进聚焦”派生列与任务十二冻结公式不一致，因此聚类只使用原始题项，不使用这两个既有派生列。",
            "",
        ]
    )
    return "\n".join(lines)


def _write_readme(summary: dict) -> None:
    text = "\n".join(
        [
            "# Task 12 relationship profile outputs",
            "",
            "本目录只保存题项映射和聚合验证结果。逐行题项矩阵保存在 `private/item_matrices.npz`，该目录被 `.gitignore` 排除，不得提交。",
            "",
            f"- 调节聚焦矩阵：{summary['matrix_shapes']['regulatory_focus']}（行×题）",
            f"- Micro YSQ-18 矩阵：{summary['matrix_shapes']['micro_ysq']}（行×题）",
            f"- 亲密关系启动意向与主动行为矩阵：{summary['matrix_shapes']['relationship']}（行×题）",
            "- 开放题原文未写入任何输出。",
            "- 画像建模只能用于探索性、非诊断、非人格定性解释。",
            "",
        ]
    )
    (OUTPUT_ROOT / "README.md").write_text(text, encoding="utf-8", newline="\n")


def main() -> int:
    for path in (ORIGINAL_PATH, ALL_DATA_PATH, CLEANED_PATH, SCORING_PATH):
        if not path.exists():
            raise SystemExit(f"missing task12 source: {path}")
    original = _read_excel_frame(ORIGINAL_PATH)
    all_data = _read_excel_frame(ALL_DATA_PATH)
    cleaned = _read_excel_frame(CLEANED_PATH)
    mapping = build_item_mapping(list(original.columns), list(all_data.columns))
    dimensions = calculate_dimensions(all_data)
    regulatory = _numeric(all_data, RF_COLUMNS)
    micro_ysq = _numeric(all_data, YSQ_COLUMNS)
    relationship = _numeric(all_data, RELATIONSHIP_COLUMNS)
    summary = {
        "schema_version": "task12-relationship-dataset-v1",
        "raw_text_included": False,
        "source_hashes": {
            path.name: _sha256(path) for path in (ORIGINAL_PATH, ALL_DATA_PATH, CLEANED_PATH, SCORING_PATH)
        },
        "sample_counts": {
            "original": int(len(original)),
            "all_data": int(len(all_data)),
            "cleaned": int(len(cleaned)),
        },
        "matrix_shapes": {
            "regulatory_focus": list(regulatory.shape),
            "micro_ysq": list(micro_ysq.shape),
            "relationship": list(relationship.shape),
        },
        "observed_ranges": {
            "regulatory_focus": _observed_range(all_data, RF_COLUMNS),
            "micro_ysq": _observed_range(all_data, YSQ_COLUMNS),
            "relationship": _observed_range(all_data, RELATIONSHIP_COLUMNS),
        },
        "within_all_data_validation": _within_source_validation(dimensions, all_data),
        "cleaned_cross_sample_validation": _cross_sample_summary(dimensions, cleaned),
        "source_conflicts": [
            "原始量表、全部数据1.0、清洗好的469份样本数不同，不能无ID逐行拼接。",
            "当前调节聚焦题项观测为1-5，新版问卷和计分说明写1-7。",
            "当前数据PBC为6题，新版问卷与信度报告按4题。",
            "全部数据1.0中的促进聚焦派生列与冻结公式不一致，但按题项复算的均值与469份清洗表PF均值接近，应以题项复算为建模输入并保留人工复核。",
            "全部数据1.0中的PBC派生列可由PBC3-PBC6四题精确复现，不是六题均值。",
        ],
    }
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    PRIVATE_ROOT.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        PRIVATE_ROOT / "item_matrices.npz",
        regulatory_focus=regulatory.to_numpy(dtype=float),
        regulatory_focus_columns=np.array(RF_COLUMNS),
        micro_ysq=micro_ysq.to_numpy(dtype=float),
        micro_ysq_columns=np.array(YSQ_COLUMNS),
        relationship=relationship.to_numpy(dtype=float),
        relationship_columns=np.array(RELATIONSHIP_COLUMNS),
    )
    pd.DataFrame(mapping).to_csv(OUTPUT_ROOT / "item_mapping_preview.csv", index=False, encoding="utf-8-sig")
    (OUTPUT_ROOT / "dimension_validation_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n"
    )
    _write_readme(summary)
    DOC_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    DOC_OUTPUT.write_text(_mapping_markdown(mapping, summary), encoding="utf-8", newline="\n")
    print(
        json.dumps(
            {
                "ok": True,
                "mapping_count": len(mapping),
                "raw_text_included": False,
                "matrix_shapes": summary["matrix_shapes"],
                "output": str(OUTPUT_ROOT),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
